import logging
import glob
import base64
import io
import numpy as np
import pandas as pd
import pdfkit
from pathlib import Path
from datetime import datetime
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt


logging.basicConfig(
    filename='sensor_pipeline.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s - %(message)s'
)

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']

# PDF Configuration
WKHTML_PATH = r'C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe'
PDF_CONFIG = pdfkit.configuration(wkhtmltopdf=WKHTML_PATH)

PDF_OPTIONS = {
    'page-size': 'A4', 'margin-top': '25mm', 'margin-bottom': '20mm',
    'margin-left': '15mm', 'margin-right': '15mm',
    'header-center': 'Confidential Corporate Report',
    'footer-left': 'Generated on: [date]', 'footer-right': 'Page [page] of [toPage]',
    'footer-font-name': 'Arial', 'footer-font-size': '9',
    'enable-local-file-access': None, 'encoding': 'UTF-8',
}


class SensorPipeline:
    def __init__(self, input_dir: str, output_dir: str, mode: str = "monthly"):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.mode = mode.lower()
        self._data = None
        self._report = None

    def run(self):
        """Main entry point for both modes"""
        if self.mode == "hourly":
            self._run_hourly()
        else:
            self._run_monthly()

    # ====================== MONTHLY ======================
    def _run_monthly(self):
        logging.info("Starting MONTHLY pipeline...")
        self._load_clean_monthly()
        self._analyze()
        self._export_excel_monthly()
        self._export_html_monthly()
        logging.info("Monthly pipeline finished successfully.")

    def _load_clean_monthly(self):
        dfs = []
        for file in glob.glob(str(self.input_dir / "*.csv")):
            df = pd.read_csv(file)
            df.columns = df.columns.str.strip().str.title().str.replace(' ', '_')
            df = df.drop_duplicates()

            df['Pressure_Hpa'] = pd.to_numeric(df.get('Pressure_Hpa'), errors='coerce').fillna(0)
            df['Status'] = df.get('Status', '').astype(str).str.strip().str.title()
            df.loc[~df['Status'].isin(['Ok', 'Warning', 'Critical']), 'Status'] = 'Unknown'

            if 'Temperature_C' in df.columns and df['Temperature_C'].dtype == object:
                temp = df['Temperature_C'].astype(str)
                value = temp.str.extract(r'([\d.]+)')[0].astype(float)
                unit = temp.str.extract(r'([CF])')[0]
                df['Temperature_C'] = np.where(unit == 'F', (value - 32) * 5 / 9, value)

            df['Source_File'] = Path(file).name
            df['Month'] = pd.to_datetime(df['Timestamp'], errors='coerce').dt.month
            dfs.append(df)

        self._data = sorted(dfs, key=lambda x: x['Month'].iloc[0] if not x.empty else 0)

    def _analyze(self):
        df_summaries, df_events, plots = [], [], []
        for df in self._data:
            # Temperature Trend Plot
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.plot(df['Temperature_C'])
            ax.set_title("Temperature Trend")
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
            buf.seek(0)
            plots.append(base64.b64encode(buf.read()).decode('utf-8'))
            plt.close(fig)

            # Summary Stats
            summary = df.groupby('Sensor_Id').agg({
                'Temperature_C': ['mean', 'max', 'min'],
                'Pressure_Hpa': ['mean', 'max', 'min'],
                'Vibration': ['mean', 'max', 'min']
            })
            summary.columns = [f"{stat} {col}" for col, stat in summary.columns]
            df_summaries.append(summary)

            df_events.append(df[df['Status'].isin(['Warning', 'Critical'])])

        self._report = df_summaries, df_events, plots

    def _export_excel_monthly(self):
        out_dir = self.output_dir / "excels"
        out_dir.mkdir(parents=True, exist_ok=True)

        summaries, events, _ = self._report
        for df, summary, event in zip(self._data, summaries, events):
            month_num = int(df['Month'].iloc[0])
            month_name = MONTHS[month_num - 1] if 1 <= month_num <= 12 else "Unknown"
            filepath = out_dir / f"Month-{month_name}-Full_Report.xlsx"

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Raw_Data', index=False)
                summary.to_excel(writer, sheet_name='Summary', index=True)
                event.to_excel(writer, sheet_name='Events', index=False)

            # Auto-adjust columns + formatting
            wb = load_workbook(filepath)
            for ws in wb.worksheets:
                ws.freeze_panes = 'A2'
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col) + 2
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max_len

            wb.save(filepath)

    def _export_html_monthly(self):
        out_dir = self.output_dir / "htmls"
        out_dir.mkdir(parents=True, exist_ok=True)
        template_path = Path("templates/report template.html")   # Adjust path as needed

        summaries, events, plots = self._report
        for df, summary, event, plot in zip(self._data, summaries, events, plots):
            month_num = int(df['Month'].iloc[0])
            month_name = MONTHS[month_num - 1] if 1 <= month_num <= 12 else "Unknown"

            with open(template_path, 'r', encoding='utf-8') as f:
                template = Template(f.read())

            html = template.render(
                month_name=month_name,
                generated_at=datetime.now(),
                sensor_data=df.to_dict('records'),
                summary_table=summary.to_html(),
                events_table=event.to_html(index=False),
                image=f"data:image/png;base64,{plot}"
            )

            (out_dir / f"{month_name}_Report.html").write_text(html, encoding='utf-8')

    # ====================== HOURLY ======================
    def _run_hourly(self):
        logging.info("Starting HOURLY pipeline...")
        self._load_clean_hourly()
        self._export_hourly_pdf()
        logging.info("Hourly report generated.")

    def _load_clean_hourly(self):
        latest_file = max(self.input_dir.glob("*.csv"), key=lambda f: f.stat().st_mtime)
        df = pd.read_csv(latest_file)

        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        df = df.drop_duplicates()

        df['pressure_hpa'] = pd.to_numeric(df.get('pressure_hpa'), errors='coerce').fillna(0)
        df['status'] = df.get('status', '').astype(str).str.strip().str.title()
        df.loc[~df['status'].isin(['Normal','Ok','Warning','Critical']), 'status'] = 'Unknown'

        if 'temperature_c' in df.columns and df['temperature_c'].dtype == object:
            temp = df['temperature_c'].astype(str)
            value = temp.str.extract(r'([\d.]+)')[0].astype(float)
            unit = temp.str.extract(r'([CF])')[0]
            df['temperature_c'] = np.where(unit == 'F', (value - 32) * 5 / 9, value)

        self._data = df

    def _export_hourly_pdf(self):
        out_dir = self.output_dir / "pdfs"
        out_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        pdf_path = out_dir / f"{timestamp}_Hourly_Report.pdf"

        template_path = Path("templates/Report.html")

        with open(template_path, 'r', encoding='utf-8') as f:
            template = Template(f.read())

        html_content = template.render(
            generated_at=datetime.now(),
            sensor_data=self._data.to_dict('records'),
            location_name="Main Plant"
        )

        pdfkit.from_string(html_content, str(pdf_path), options=PDF_OPTIONS, configuration=PDF_CONFIG)
        print(f"✅ Hourly PDF generated: {pdf_path}")